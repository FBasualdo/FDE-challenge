"""Build .xlsx binaries for dashboard exports.

The helper takes a column spec (a list of `(field_key, header_label)` pairs)
plus an iterable of dict rows and returns the workbook as bytes. Callers
stream the bytes back through FastAPI with a `Content-Disposition` header.

Design choices:
  - Bold header row + frozen pane on row 1 so the export is usable in Excel
    even at 10k rows.
  - Autofilter on the data range — quick triage in the dashboard download.
  - Column widths derived from a single scan over the data (capped at 60
    chars to keep the workbook from blowing out horizontally on long
    fields like transcripts).
  - Datetimes are written as native Excel datetimes (openpyxl converts
    `datetime` automatically); dates as date objects; everything else is
    str()'d defensively so JSONB blobs and Decimals don't crash the writer.
"""

from __future__ import annotations

from datetime import date, datetime, timezone
from decimal import Decimal
from io import BytesIO
from typing import Any, Iterable

from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

EXCEL_MEDIA_TYPE = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"

# Hard ceiling on rows per export. Anything above forces the dashboard to
# refine filters first instead of generating a multi-MB workbook on a
# request thread. Excel handles 10k rows comfortably.
EXPORT_ROW_CAP = 10_000

# Per-column auto-fit cap. Widening past this hurts more than it helps
# (long transcripts, long JSON blobs). Excel users can still resize manually.
_MAX_COL_WIDTH = 60
_MIN_COL_WIDTH = 10


def build_filename(prefix: str, *, now: datetime | None = None) -> str:
    """Return a timestamped filename like 'calls-20260501-143205.xlsx'."""
    ts = (now or datetime.now(timezone.utc)).strftime("%Y%m%d-%H%M%S")
    return f"{prefix}-{ts}.xlsx"


def _coerce_cell(value: Any) -> Any:
    """Translate Python types into something openpyxl handles cleanly.

    - `datetime`: Excel doesn't support tz-aware datetimes, so we
      normalize to UTC and drop the tzinfo. The cell still renders as
      an Excel datetime; the user reads UTC, which matches every other
      surface in the dashboard.
    - `date`: passes through.
    - `Decimal`: becomes float (openpyxl special-cases Decimal but
      float is friendlier in spreadsheet pivots).
    - `bool`: stays bool — Excel renders them as TRUE/FALSE.
    - `dict` / `list`: JSON-serialized so structured columns (e.g.
      `raw_response`) land as a single readable cell instead of
      crashing the writer with "unsupported type".
    - `None`: becomes empty string so the cell renders blank, not "None".
    """
    if value is None:
        return ""
    if isinstance(value, datetime):
        if value.tzinfo is not None:
            value = value.astimezone(timezone.utc).replace(tzinfo=None)
        return value
    if isinstance(value, (date, bool, int, float, str)):
        return value
    if isinstance(value, Decimal):
        return float(value)
    if isinstance(value, (dict, list)):
        # Compact JSON keeps the cell on one line — important for autofilter.
        import json

        return json.dumps(value, default=str, separators=(",", ":"))
    return str(value)


def _column_width(values: list[Any], header: str) -> float:
    """Pick a reasonable column width based on the longest rendered value."""
    longest = len(str(header))
    for v in values:
        if v is None:
            continue
        # Datetimes render at ~19 chars in Excel; cap their measured length
        # so they don't blow widths out of proportion.
        rendered = str(v) if not isinstance(v, (datetime, date)) else "0000-00-00 00:00:00"
        longest = max(longest, len(rendered))
    return max(_MIN_COL_WIDTH, min(_MAX_COL_WIDTH, longest + 2))


def write_workbook(
    sheet_name: str,
    columns: list[tuple[str, str]],
    rows: Iterable[dict[str, Any]],
) -> bytes:
    """Build an .xlsx with a header row + data rows, return the binary.

    Args:
      sheet_name: name shown on the sheet tab.
      columns: ordered `(field_key, header_label)` pairs. `field_key` is
        the dict key in each row; `header_label` is what the user sees.
      rows: iterable of dicts. Missing keys render as empty cells.
    """
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name

    field_keys = [c[0] for c in columns]
    headers = [c[1] for c in columns]

    # Header row.
    ws.append(headers)
    bold = Font(bold=True)
    for col_idx in range(1, len(headers) + 1):
        ws.cell(row=1, column=col_idx).font = bold

    # Data rows. Materialize once so we can also auto-fit widths without
    # iterating the source twice.
    materialized: list[list[Any]] = []
    for raw in rows:
        row_values = [_coerce_cell(raw.get(k)) for k in field_keys]
        materialized.append(row_values)
        ws.append(row_values)

    # Frozen pane below the header so it stays visible while scrolling.
    ws.freeze_panes = "A2"

    # Autofilter across the populated range. With zero data rows, anchor
    # to the header alone so Excel still shows the filter dropdowns.
    last_row = max(1, len(materialized) + 1)
    last_col_letter = get_column_letter(len(headers))
    ws.auto_filter.ref = f"A1:{last_col_letter}{last_row}"

    # Auto-fit column widths from the materialized data (single pass).
    for idx, (_, header) in enumerate(columns):
        col_values = [row[idx] for row in materialized]
        ws.column_dimensions[get_column_letter(idx + 1)].width = _column_width(col_values, header)

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()
